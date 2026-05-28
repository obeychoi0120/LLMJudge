import pandas as pd
import json
import os
from openpyxl.utils import get_column_letter
from utils import load_jsonl, load_content_indices

content_indices = load_content_indices()

_MODE_ORDER = ["blank", "meta", "video", "kss", "raw", "raw_with_mmvlm", "imgvlm_chunk2", "imgvlm_graph"]

def _sort_modes(modes):
    """MODE_ORDER 순서대로 정렬, 목록에 없는 모드는 뒤에 알파벳순."""
    ordered = [m for m in _MODE_ORDER if m in modes]
    ordered += sorted(set(modes) - set(_MODE_ORDER))
    return ordered

def aggregate_vh_scores(input_dir, output_dir):
    """VH Score를 query_type별로 따로 집계하여 JSON으로 저장합니다."""
    scores_file = os.path.join(input_dir, "voice_hint_scores.jsonl")
    output_file = os.path.join(output_dir, "voice_hint_scores_aggregated.json")

    if not os.path.exists(scores_file):
        print(f"[Skip] {scores_file} 파일이 없어 VH Aggregation을 건너뜁니다.")
        return

    print(f"Aggregating VH scores from {scores_file}...")
    try:
        data = load_jsonl(scores_file)
    except Exception as e:
        print(f"Error: Failed to parse {scores_file}: {e}")
        return

    # query_type별 메트릭 정의
    _METRICS_BY_TYPE = {
        "content_anchored": ["temporal_immersion", "content_depth", "total_score"],
        "tangential": ["temporal_immersion", "curiosity_and_hook", "total_score"],
    }

    # query_type별 독립 집계
    # by_query_type[q_type]["by_video"][content_id][mode][metric] = [values...]
    # by_query_type[q_type]["overall"][mode][metric] = [values...]
    agg = {}

    for record in data:
        c_id = record.get("content_id")
        if not c_id:
            continue

        m = record.get("mode", "unknown")
        q_type = record.get("query_type", "tangential")
        judge_data = record.get("judge", {})
        ts = record.get("total_score")

        metrics = _METRICS_BY_TYPE.get(q_type, _METRICS_BY_TYPE["tangential"])

        if q_type not in agg:
            agg[q_type] = {"by_video_raw": {}, "overall_raw": {}}

        by_video_raw = agg[q_type]["by_video_raw"]
        overall_raw = agg[q_type]["overall_raw"]

        if c_id not in by_video_raw:
            by_video_raw[c_id] = {}
        if m not in by_video_raw[c_id]:
            by_video_raw[c_id][m] = {met: [] for met in metrics}
        if m not in overall_raw:
            overall_raw[m] = {met: [] for met in metrics}

        # 개별 메트릭 수집 (total_score 제외)
        for met in metrics[:-1]:
            if met in judge_data:
                val = judge_data[met].get("score")
                if isinstance(val, (int, float)):
                    by_video_raw[c_id][m][met].append(val)
                    overall_raw[m][met].append(val)

        if isinstance(ts, (int, float)):
            by_video_raw[c_id][m]["total_score"].append(ts)
            overall_raw[m]["total_score"].append(ts)

    # 평균 계산
    final_output = {}
    for q_type, raw_data in agg.items():
        metrics = _METRICS_BY_TYPE.get(q_type, _METRICS_BY_TYPE["tangential"])

        # content_id별 평균
        results_by_video = {}
        for c_id, modes_data in raw_data["by_video_raw"].items():
            avg_video = {}
            for m, raw in modes_data.items():
                avg_mode = {}
                for met in metrics:
                    s_list = raw.get(met, [])
                    if s_list:
                        avg_mode[met] = round(sum(s_list) / len(s_list), 2)
                if avg_mode:
                    avg_video[m] = avg_mode
            if avg_video:
                results_by_video[c_id] = avg_video

        # 전체 평균
        results_overall = {}
        for m, raw in raw_data["overall_raw"].items():
            avg_mode = {}
            for met in metrics:
                s_list = raw.get(met, [])
                if s_list:
                    avg_mode[met] = round(sum(s_list) / len(s_list), 2)
            if avg_mode:
                results_overall[m] = avg_mode

        final_output[q_type] = {"by_video": results_by_video, "overall": results_overall}

    with open(output_file, "w", encoding="utf-8") as f:
        json.dump(final_output, f, indent=4, ensure_ascii=False)
    print(f"[Done] VH Score Aggregation (by query_type): {output_file}")

def export_vh_details(input_dir, output_dir):
    """Voice Hint Judge 상세 결과를 Excel로 내보냅니다."""
    vh_scores_path = os.path.join(input_dir, "voice_hint_scores.jsonl")
    if not os.path.exists(vh_scores_path):
        print(f"[Skip] {vh_scores_path} 파일이 없어 VH Details 내보내기를 건너뜁니다.")
        return

    vh_data = load_jsonl(vh_scores_path)

    flat_vh = []
    # 모든 가능한 메트릭 키
    all_metrics = ["temporal_immersion", "content_depth", "curiosity_and_hook"]

    for item in vh_data:
        cid = item.get("content_id", "")
        item_scene_idx = item.get("scene_idx")

        if "query" in item and "judge" in item:
            judge = item.get("judge", {})
            total = item.get("total_score", "")
            q_type = item.get("query_type", "")
            row = {
                "index": content_indices.get(cid, 999),
                "content_id": cid,
                "scene_idx": item_scene_idx,
                "mode": item.get("mode", ""),
                "query_type": q_type,
                "query": item.get("query", ""),
            }
            for met in all_metrics:
                met_data = judge.get(met, {})
                row[f"rationale_{met}"] = met_data.get("rationale", "") if isinstance(met_data, dict) else ""
                row[met] = met_data.get("score", "") if isinstance(met_data, dict) else ""
            row["total_score"] = total
            flat_vh.append(row)

    if not flat_vh:
        print("[Skip] Voice Hint Score 데이터가 비어 있습니다.")
        return

    # index 순으로 정렬
    flat_vh.sort(key=lambda x: (
        x.get("index", 999),
        x.get("content_id", ""),
        x.get("scene_idx", 0),
        x.get("mode", ""),
        x.get("query", ""),
    ))

    df = pd.DataFrame(flat_vh)
    out_path = os.path.join(output_dir, "vh_score_details.xlsx")

    writer = pd.ExcelWriter(out_path, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name='VH Score Details')
    worksheet = writer.sheets['VH Score Details']
    
    col_widths = {
        "index": 10,
        "content_id": 20, "scene_idx": 10, "mode": 10, "query_type": 16, "query": 40,
        "rationale_temporal_immersion": 40, "temporal_immersion": 12,
        "rationale_content_depth": 40, "content_depth": 12,
        "rationale_curiosity_and_hook": 40, "curiosity_and_hook": 12,
        "total_score": 12
    }
    for idx, col in enumerate(df.columns):
        width = col_widths.get(col, 15)
        col_letter = get_column_letter(idx + 1)
        worksheet.column_dimensions[col_letter].width = width
    writer.close()

    print(f"Created: {out_path}")

def _write_pivot_scores_xlsx(out_path, sheet_name, metrics, ordered_modes, all_cids, by_video, overall_data):
    """메트릭 × 모드 피벗 구조의 Excel을 openpyxl로 직접 작성합니다.

    헤더 구조 (2행):
      Row 1: content_id (merged) | metric1 (merged across modes) | metric2 (merged) | ...
      Row 2:                     | mode1 | mode2 | ...           | mode1 | mode2 | ...  | ...
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    n_modes = len(ordered_modes)

    # ── 스타일 정의 ──
    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    metric_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # ── Row 1: 상위 헤더 (index + content_id + 메트릭 이름, 각 메트릭은 n_modes 열 병합) ──
    # index (col 1)
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    cell = ws.cell(row=1, column=1, value="index")
    cell.font = header_font
    cell.alignment = center_align
    cell.border = thin_border

    # content_id (col 2)
    ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
    cell = ws.cell(row=1, column=2, value="content_id")
    cell.font = header_font
    cell.alignment = center_align
    cell.border = thin_border

    col = 3  # 현재 열 위치 (1-indexed)
    for met in metrics:
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + n_modes - 1)
        cell = ws.cell(row=1, column=col, value=met)
        cell.font = header_font
        cell.alignment = center_align
        cell.fill = metric_fill
        cell.border = thin_border
        col += n_modes

    # ── Row 2: 하위 헤더 (모드 이름 반복) ──
    col = 3
    for _ in metrics:
        for mode in ordered_modes:
            cell = ws.cell(row=2, column=col, value=mode)
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
            col += 1

    # ── Data rows (Row 3~) ──
    data_start_row = 3
    sorted_cids = sorted(all_cids, key=lambda c: content_indices.get(c, 999))
    for r_idx, cid in enumerate(sorted_cids):
        row_num = data_start_row + r_idx
        idx_val = content_indices.get(cid, 999)
        ws.cell(row=row_num, column=1, value=idx_val).border = thin_border
        ws.cell(row=row_num, column=2, value=cid).border = thin_border
        col = 3
        modes_data = by_video.get(cid, {})
        for met in metrics:
            for mode in ordered_modes:
                val = modes_data.get(mode, {}).get(met, "")
                c = ws.cell(row=row_num, column=col, value=val if val != "" else None)
                c.border = thin_border
                col += 1

    # OVERALL 행
    if overall_data:
        row_num = data_start_row + len(all_cids)
        ws.cell(row=row_num, column=1, value="").border = thin_border
        c = ws.cell(row=row_num, column=2, value="OVERALL")
        c.font = Font(bold=True)
        c.border = thin_border
        col = 3
        for met in metrics:
            for mode in ordered_modes:
                val = overall_data.get(mode, {}).get(met, "")
                c = ws.cell(row=row_num, column=col, value=val if val != "" else None)
                c.border = thin_border
                col += 1

    # ── 열 너비 조정 ──
    ws.column_dimensions[get_column_letter(1)].width = 10
    ws.column_dimensions[get_column_letter(2)].width = 28
    total_cols = 2 + len(metrics) * n_modes
    for ci in range(3, total_cols + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 14

    wb.save(out_path)
    print(f"Created: {out_path}")


def _write_response_pivot_scores_xlsx(out_path, sheet_name, metrics, track_config, all_cids, agg):
    """B-Track (VH Response)용 3행 헤더 피벗 Excel을 openpyxl로 작성합니다.

    헤더 구조 (3행):
      Row 1: content_id (merged v) | metric1 (merged h across 8 cols) | ...
      Row 2:                       | high-context (merged h 4 cols) | low-context (merged h 4 cols) | ...
      Row 3:                       | blank | video | ...            | blank | imgvlm_sentence | ... | ...
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # ── 스타일 정의 ──
    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    metric_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    track_fills = {
        "high-context": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        "low-context": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    }

    # ── Row 1-3: index 및 content_id (세로 병합) ──
    # index (col 1)
    ws.merge_cells(start_row=1, start_column=1, end_row=3, end_column=1)
    cell = ws.cell(row=1, column=1, value="index")
    cell.font = header_font
    cell.alignment = center_align
    cell.border = thin_border

    # content_id (col 2)
    ws.merge_cells(start_row=1, start_column=2, end_row=3, end_column=2)
    cell = ws.cell(row=1, column=2, value="content_id")
    cell.font = header_font
    cell.alignment = center_align
    cell.border = thin_border

    # 격자 테두리 초기화
    for r in range(1, 4):
        ws.cell(row=r, column=1).border = thin_border
        ws.cell(row=r, column=2).border = thin_border

    col = 3  # 현재 열 위치 (1-indexed)
    for met in metrics:
        total_modes_for_metric = sum(len(modes) for modes in track_config.values())
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + total_modes_for_metric - 1)
        cell = ws.cell(row=1, column=col, value=met)
        cell.font = header_font
        cell.alignment = center_align
        cell.fill = metric_fill
        
        for c in range(col, col + total_modes_for_metric):
            ws.cell(row=1, column=c).border = thin_border

        sub_col = col
        for track_name, modes in track_config.items():
            n_track_modes = len(modes)
            ws.merge_cells(start_row=2, start_column=sub_col, end_row=2, end_column=sub_col + n_track_modes - 1)
            track_cell = ws.cell(row=2, column=sub_col, value=track_name)
            track_cell.font = header_font
            track_cell.alignment = center_align
            track_cell.fill = track_fills.get(track_name, metric_fill)
            
            for c in range(sub_col, sub_col + n_track_modes):
                ws.cell(row=2, column=c).border = thin_border

            for mode in modes:
                mode_cell = ws.cell(row=3, column=sub_col, value=mode)
                mode_cell.font = header_font
                mode_cell.alignment = center_align
                mode_cell.border = thin_border
                sub_col += 1

        col += total_modes_for_metric

    # ── Data rows (Row 4~) ──
    data_start_row = 4
    sorted_cids = sorted(all_cids, key=lambda c: content_indices.get(c, 999))
    for r_idx, cid in enumerate(sorted_cids):
        row_num = data_start_row + r_idx
        idx_val = content_indices.get(cid, 999)
        ws.cell(row=row_num, column=1, value=idx_val).border = thin_border
        ws.cell(row=row_num, column=2, value=cid).border = thin_border
        
        col = 3
        for met in metrics:
            for track_name, modes in track_config.items():
                for mode in modes:
                    val = ""
                    for q_type, q_type_data in agg.items():
                        by_video = q_type_data.get("by_video", {})
                        modes_data = by_video.get(cid, {})
                        if mode in modes_data and met in modes_data[mode]:
                            val = modes_data[mode][met]
                            break
                    c = ws.cell(row=row_num, column=col, value=val if val != "" else None)
                    c.border = thin_border
                    col += 1

    # OVERALL 행
    if all_cids:
        row_num = data_start_row + len(all_cids)
        ws.cell(row=row_num, column=1, value="").border = thin_border
        c = ws.cell(row=row_num, column=2, value="OVERALL")
        c.font = Font(bold=True)
        c.border = thin_border
        
        col = 3
        for met in metrics:
            for track_name, modes in track_config.items():
                for mode in modes:
                    val = ""
                    for q_type, q_type_data in agg.items():
                        overall_data = q_type_data.get("overall", {})
                        if mode in overall_data and met in overall_data[mode]:
                            val = overall_data[mode][met]
                            break
                    c = ws.cell(row=row_num, column=col, value=val if val != "" else None)
                    c.border = thin_border
                    col += 1

    # ── 열 너비 조정 ──
    ws.column_dimensions[get_column_letter(1)].width = 10
    ws.column_dimensions[get_column_letter(2)].width = 28
    total_cols = 2 + len(metrics) * sum(len(modes) for modes in track_config.values())
    for ci in range(3, total_cols + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 14

    wb.save(out_path)
    print(f"Created: {out_path}")


def export_vh_scores(input_dir, output_dir):
    """Voice Hint Judge 집계 점수를 고맥락/저맥락별 엑셀 파일로 내보냅니다."""
    agg_path = os.path.join(output_dir, "voice_hint_scores_aggregated.json")
    if not os.path.exists(agg_path):
        print(f"[Skip] {agg_path} 파일이 없어 VH Scores 내보내기를 건너뜁니다.")
        return

    with open(agg_path, "r", encoding="utf-8") as f:
        agg = json.load(f)

    # query_type별 메트릭 정의
    _METRICS_BY_TYPE = {
        "content_anchored": ["temporal_immersion", "content_depth", "total_score"],
        "tangential": ["temporal_immersion", "curiosity_and_hook", "total_score"],
    }

    # query_type별로 별도 Excel 파일 생성
    for q_type, q_type_data in agg.items():
        metrics = _METRICS_BY_TYPE.get(q_type, ["total_score"])
        by_video = q_type_data.get("by_video", {})
        overall = q_type_data.get("overall", {})

        all_cids = list(by_video.keys())
        all_modes = set()
        for modes_data in by_video.values():
            all_modes.update(modes_data.keys())
        all_modes.update(overall.keys())
        ordered_modes = _sort_modes(all_modes)

        # 트랙에 해당하는 모드만 필터링
        target_modes = ["video", "kss", "raw", "raw_with_mmvlm"] if q_type == "content_anchored" else ["meta", "imgvlm_sentence", "imgvlm_chunk2", "imgvlm_graph"]
        ordered_modes = [m for m in ordered_modes if m in target_modes]

        if not ordered_modes:
            continue

        suffix = "high_context" if q_type == "content_anchored" else "low_context"
        out_path = os.path.join(output_dir, f"vh_scores_{suffix}.xlsx")
        sheet_name = "VH High-Context" if q_type == "content_anchored" else "VH Low-Context"
        _write_pivot_scores_xlsx(
            out_path, sheet_name, metrics, ordered_modes,
            all_cids, by_video, overall,
        )



def aggregate_vh_response_scores(input_dir, output_dir, query_source="kss"):
    """VH Response Score를 query_type별로 따로 집계하여 JSON으로 저장합니다."""
    scores_file = os.path.join(input_dir, f"vh_response_scores_{query_source}.jsonl")
    output_file = os.path.join(output_dir, f"vh_response_scores_aggregated_{query_source}.json")

    if not os.path.exists(scores_file) and query_source == "kss":
        fallback_file = os.path.join(input_dir, "vh_response_scores.jsonl")
        if os.path.exists(fallback_file):
            scores_file = fallback_file
            output_file = os.path.join(output_dir, "vh_response_scores_aggregated.json")

    if not os.path.exists(scores_file):
        print(f"[Skip] {scores_file} 파일이 없어 VH Response Score Aggregation을 건너뜁니다.")
        return

    print(f"Aggregating VH Response scores from {scores_file}...")
    try:
        data = load_jsonl(scores_file)
    except Exception as e:
        print(f"Error: Failed to parse {scores_file}: {e}")
        return

    metrics = ["answer_relevance", "factual_precision", "informativeness", "total_score"]

    # query_type별 독립 집계
    agg = {}

    for record in data:
        c_id = record.get("content_id")
        mode = record.get("mode")
        q_type = record.get("query_type", "unknown")
        judge = record.get("judge", {})
        if not c_id or not mode or not judge:
            continue

        # total_score: 3개 metric 합계
        ts = sum(
            judge.get(k, {}).get("score", 0)
            for k in metrics[:-1]
            if isinstance(judge.get(k), dict)
        )

        if q_type not in agg:
            agg[q_type] = {"by_video_raw": {}, "overall_raw": {}}

        by_video_raw = agg[q_type]["by_video_raw"]
        overall_raw = agg[q_type]["overall_raw"]

        if c_id not in by_video_raw:
            by_video_raw[c_id] = {}
        if mode not in by_video_raw[c_id]:
            by_video_raw[c_id][mode] = {met: [] for met in metrics}
        if mode not in overall_raw:
            overall_raw[mode] = {met: [] for met in metrics}

        for met in metrics[:-1]:
            val = judge.get(met, {}).get("score") if isinstance(judge.get(met), dict) else None
            if isinstance(val, (int, float)):
                by_video_raw[c_id][mode][met].append(val)
                overall_raw[mode][met].append(val)

        by_video_raw[c_id][mode]["total_score"].append(ts)
        overall_raw[mode]["total_score"].append(ts)

    # 평균 계산
    final_output = {}
    for q_type, raw_data in agg.items():
        results_by_video = {}
        for c_id, modes_data in raw_data["by_video_raw"].items():
            avg_video = {}
            for mode, raw in modes_data.items():
                avg_mode = {}
                for met in metrics:
                    s_list = raw.get(met, [])
                    if s_list:
                        avg_mode[met] = round(sum(s_list) / len(s_list), 2)
                if avg_mode:
                    avg_video[mode] = avg_mode
            if avg_video:
                results_by_video[c_id] = avg_video

        results_overall = {}
        for mode, raw in raw_data["overall_raw"].items():
            avg_mode = {}
            for met in metrics:
                s_list = raw.get(met, [])
                if s_list:
                    avg_mode[met] = round(sum(s_list) / len(s_list), 2)
            if avg_mode:
                results_overall[mode] = avg_mode

        final_output[q_type] = {"by_video": results_by_video, "overall": results_overall}

    with open(output_file, "w", encoding="utf-8") as f:
        json.dump(final_output, f, indent=4, ensure_ascii=False)
    print(f"[Done] VH Response Score Aggregation (by query_type): {output_file}")


def export_vh_response_details(input_dir, output_dir, query_source="kss"):
    """VH Response Judge 상세 결과를 Excel로 내보냅니다."""
    scores_path = os.path.join(input_dir, f"vh_response_scores_{query_source}.jsonl")
    responses_path = os.path.join(input_dir, f"vh_responses_{query_source}.jsonl")
    out_path = os.path.join(output_dir, f"vh_response_score_details_{query_source}.xlsx")

    if not os.path.exists(scores_path) and query_source == "kss":
        fallback_scores = os.path.join(input_dir, "vh_response_scores.jsonl")
        fallback_responses = os.path.join(input_dir, "vh_responses.jsonl")
        if os.path.exists(fallback_scores):
            scores_path = fallback_scores
            responses_path = fallback_responses
            out_path = os.path.join(output_dir, "vh_response_score_details.xlsx")

    if not os.path.exists(scores_path):
        print(f"[Skip] {scores_path} 파일이 없어 VH Response Details 내보내기를 건너뜁니다.")
        return

    # (content_id, scene_idx, mode, query) → response 매핑
    answer_map = {}
    if os.path.exists(responses_path):
        for r in load_jsonl(responses_path):
            c = r.get("content_id")
            s = r.get("scene_idx")
            m = r.get("mode")
            q = r.get("query")
            ans = r.get("answer")
            if c and s is not None and m and q:
                answer_map[(c, s, m, q)] = ans

    data = load_jsonl(scores_path)
    score_keys = ["answer_relevance", "factual_precision", "informativeness"]

    flat_rows = []
    for item in data:
        c_id    = item.get("content_id", "")
        s_idx   = item.get("scene_idx")
        mode    = item.get("mode", "")
        query   = item.get("query", "")
        judge   = item.get("judge", {})

        if not isinstance(judge, dict):
            continue

        ts = sum(
            judge.get(k, {}).get("score", 0)
            for k in score_keys
            if isinstance(judge.get(k), dict)
        )
        response_text = answer_map.get((c_id, s_idx, mode, query), "")
        q_type = item.get("query_type", "")
        row = {"index": content_indices.get(c_id, 999), "content_id": c_id, "scene_idx": s_idx, "mode": mode, "query_type": q_type, "query": query, "response": response_text}
        
        for k in score_keys:
            met_data = judge.get(k, {})
            row[f"rationale_{k}"] = met_data.get("rationale", "") if isinstance(met_data, dict) else ""
            row[k]               = met_data.get("score", "")     if isinstance(met_data, dict) else ""
        row["total_score"] = ts
        flat_rows.append(row)

    if not flat_rows:
        print("[Skip] VH Response Score 데이터가 비어 있습니다.")
        return

    # index 순으로 정렬
    flat_rows.sort(key=lambda x: (
        x.get("index", 999),
        x.get("content_id", ""),
        x.get("scene_idx", 0),
        x.get("mode", ""),
        x.get("query", ""),
    ))

    df = pd.DataFrame(flat_rows)

    writer = pd.ExcelWriter(out_path, engine="openpyxl")
    df.to_excel(writer, index=False, sheet_name="VH Resp Score Details")
    worksheet = writer.sheets["VH Resp Score Details"]
    col_widths = {
        "index": 10,
        "content_id": 22, "scene_idx": 10, "mode": 12, "query_type": 16, "query": 40, "response": 60,
        "rationale_answer_relevance": 45,  "answer_relevance": 16,
        "rationale_factual_precision": 45, "factual_precision": 16,
        "rationale_informativeness": 45,  "informativeness": 16,
        "total_score": 12,
    }
    for idx, col in enumerate(df.columns):
        width = col_widths.get(col, 15)
        col_letter = get_column_letter(idx + 1)
        worksheet.column_dimensions[col_letter].width = width
    writer.close()
    print(f"Created: {out_path}")


def export_vh_response_scores(input_dir, output_dir, query_source="kss"):
    """VH Response Judge 집계 점수를 단일 Excel 파일(vh_response_scores_{query_source}.xlsx)로 내보냅니다.
    'high-context'와 'low-context' subcolumn을 포함합니다.
    """
    agg_path = os.path.join(output_dir, f"vh_response_scores_aggregated_{query_source}.json")
    out_path = os.path.join(output_dir, f"vh_response_scores_{query_source}.xlsx")

    if not os.path.exists(agg_path) and query_source == "kss":
        fallback_agg = os.path.join(output_dir, "vh_response_scores_aggregated.json")
        if os.path.exists(fallback_agg):
            agg_path = fallback_agg
            out_path = os.path.join(output_dir, "vh_response_scores.xlsx")

    if not os.path.exists(agg_path):
        print(f"[Skip] {agg_path} 파일이 없어 VH Response Scores 내보내기를 건너뜁니다.")
        return

    with open(agg_path, "r", encoding="utf-8") as f:
        agg = json.load(f)

    metrics = ["answer_relevance", "factual_precision", "informativeness", "total_score"]

    # content_id 전체 합집합 구하기
    all_cids = set()
    for q_type in ["content_anchored", "tangential"]:
        if q_type in agg:
            all_cids.update(agg[q_type].get("by_video", {}).keys())

    track_config = {
        "high-context": ["video", "raw", "raw_with_mmvlm"],
        "low-context": ["blank", "imgvlm_sentence", "imgvlm_chunk2", "imgvlm_graph"]
    }
    if query_source == "sourcewise":
        track_config["high-context"] = [m for m in track_config["high-context"] if m != "blank"]
        track_config["low-context"] = [m for m in track_config["low-context"] if m != "blank"]

    sheet_name = "VH Response Scores"
    
    _write_response_pivot_scores_xlsx(
        out_path, sheet_name, metrics, track_config,
        list(all_cids), agg
    )


def export_voice_hints(input_dir, output_dir):
    """Voice Hint 질문을 content_id / scene_idx 기준으로 정리하여 Excel로 내보냅니다.

    컬럼 구조:
    - content_id : 동일 content_id를 가진 첫 번째 행에만 표시, 나머지 행은 빈 칸
    - scene_idx  : 각 scene마다 표시
    - queries_{mode} : 각 mode별 "1. ~~~\n2. ~~~" 형태로 한 셀에 번호 붙여 기록
    """
    vh_path = os.path.join(input_dir, "voice_hint.jsonl")
    if not os.path.exists(vh_path):
        print(f"[Skip] {vh_path} 파일이 없어 Voice Hint Export를 건너뜁니다.")
        return

    from collections import defaultdict
    from openpyxl.styles import Alignment

    # {content_id: {scene_idx: {mode: queries_text}}}
    scenes_by_content = defaultdict(lambda: defaultdict(dict))
    found_modes = set()

    for rec in load_jsonl(vh_path):
        c_id  = rec.get("content_id")
        s_idx = rec.get("scene_idx")
        mode  = rec.get("mode", "")
        qs    = rec.get("queries", [])
        if not c_id or s_idx is None or not mode:
            continue
        numbered = "\n".join(f"{i+1}. {q}" for i, q in enumerate(qs))
        scenes_by_content[c_id][s_idx][mode] = numbered
        found_modes.add(mode)

    if not scenes_by_content:
        print("[Skip] Voice Hint 데이터가 비어 있습니다.")
        return

    ordered_modes = _sort_modes(found_modes)

    query_columns = [f"queries_{m}" for m in ordered_modes]
    
    flat_rows = []
    sorted_cids = sorted(scenes_by_content.keys(), key=lambda c: content_indices.get(c, 999))
    for c_id in sorted_cids:
        first = True
        for s_idx in sorted(scenes_by_content[c_id].keys()):
            idx_val = content_indices.get(c_id, 999)
            row = {
                "index": idx_val if first else "",
                "content_id": c_id if first else "",
                "scene_idx":  s_idx,
            }
            for mode in ordered_modes:
                row[f"queries_{mode}"] = scenes_by_content[c_id][s_idx].get(mode, "")
            flat_rows.append(row)
            first = False

    df = pd.DataFrame(flat_rows)
    out_path = os.path.join(output_dir, "voice_hints.xlsx")

    writer = pd.ExcelWriter(out_path, engine="openpyxl")
    df.to_excel(writer, index=False, sheet_name="Voice Hints")
    worksheet = writer.sheets["Voice Hints"]

    col_widths = {"index": 10, "content_id": 28, "scene_idx": 10}
    for qc in query_columns:
        col_widths[qc] = 60

    for idx, col in enumerate(df.columns):
        col_letter = get_column_letter(idx + 1)
        worksheet.column_dimensions[col_letter].width = col_widths.get(col, 20)

    query_col_letters = set()
    for qc in query_columns:
        if qc in df.columns:
            query_col_letters.add(get_column_letter(list(df.columns).index(qc) + 1))

    for row_cells in worksheet.iter_rows(min_row=2):
        for cell in row_cells:
            if cell.column_letter in query_col_letters:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(vertical="top")

    writer.close()
    print(f"Created: {out_path}")


if __name__ == "__main__":
    assets_dir = "assets"
    results_dir = "results"

    if not os.path.exists(results_dir):
        os.makedirs(results_dir)

    # assets 디렉토리 내 파일을 스캔하여 존재하는 query_source 모드 검색
    query_sources = set()
    if os.path.exists(assets_dir):
        for filename in os.listdir(assets_dir):
            if filename.startswith("vh_response_scores_") and filename.endswith(".jsonl"):
                qs = filename[len("vh_response_scores_"):-len(".jsonl")]
                query_sources.add(qs)
        if os.path.exists(os.path.join(assets_dir, "vh_response_scores.jsonl")):
            query_sources.add("kss")

    if not query_sources:
        query_sources.add("kss")

    print("="*60)
    print("1. Data Aggregation Phase")
    print("="*60)
    aggregate_vh_scores(assets_dir, results_dir)
    for qs in sorted(query_sources):
        aggregate_vh_response_scores(assets_dir, results_dir, query_source=qs)

    print("\n" + "="*60)
    print("2. Excel Export Phase")
    print("="*60)
    export_vh_details(assets_dir, results_dir)
    export_vh_scores(assets_dir, results_dir)
    for qs in sorted(query_sources):
        export_vh_response_details(assets_dir, results_dir, query_source=qs)
        export_vh_response_scores(assets_dir, results_dir, query_source=qs)
    export_voice_hints(assets_dir, results_dir)

    print("\nAll pipeline tasks completed.")